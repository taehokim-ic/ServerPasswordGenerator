from passGen import Password, PasswordList
from openpyxl import load_workbook
from openpyxl.styles import Font

passwdlist = PasswordList()
passwdlist.setRows(10)
passwdlist.setColumns(3)
passwdlist.passwordListGenerator()


wb = load_workbook('password.xlsx')
ws = wb.active
ft = Font(name='Courier New', size=11)


for i in range(passwdlist.rows):
    for j in range(passwdlist.columns):
        cell = chr(65+j) + str(i+1)   
        ws[cell].font = ft
        ws[cell] = passwdlist.passwordList[i][j]

wb.save('password.xlsx')
