import random
from openpyxl import load_workbook
from openpyxl.styles import Font

class Password():

    def __init__(self):
        
        self.password = ""

    def lessThanThree(self, passChr):

        try:
            index = self.password.index(passChr)

        except ValueError as e:
            return True

        try:
            index = self.password.index(passChr, index)
        
        except ValueError as e:
            return True

        return False

    def passwordGenerator(self, length):

        for i in range(length):
            passChr = chr(random.randint(33, 126))
            while (i == 0 and passChr == '='):
                passChr = chr(random.randint(33, 126))
            while not (self.lessThanThree(passChr)):
                passChr = chr(random.randint(33, 126))
            self.password += passChr

        return self.password

    def passwordClear(self):

        self.password = ""

class PasswordList():

    def __init__(self):
        
        self.passwordList = []
        self.rows = 0
        self.columns = 0
        self.password = Password()

    def setRows(self, row):
        self.rows = row
    
    def setColumns(self, column):
        self.columns = column

    def passwordListGenerator(self):
        
        passwords = []

        for i in range(self.rows * self.columns):
            self.password.passwordGenerator(9)
            while (self.password.password in passwords):
                self.password.passwordGenerator(9)
            passwords.append(self.password.password)
            self.password.passwordClear()
        
        for i in range(self.rows):
            _ = []
            for j in range(self.columns):
                _.append(passwords.pop())
            self.passwordList.append(_)

        return self.passwordList

passwdlist = PasswordList()
passwdlist.setRows(10)
passwdlist.setColumns(3)
passwdlist.passwordListGenerator()


wb = load_workbook('password.xlsx')
ws = wb.active
ft = Font(name='Courier New', size=11)


for i in range(passwdlist.rows):
    for j in range(passwdlist.columns):
        cell = chr(65+j) + str(i+1)   
        ws[cell].font = ft
        ws[cell] = passwdlist.passwordList[i][j]

wb.save('password.xlsx')
